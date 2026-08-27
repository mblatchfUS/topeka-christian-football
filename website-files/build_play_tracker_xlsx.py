#!/usr/bin/env python3
"""Generate the printable/editable play-count & rotation sheet as .xlsx.
Mirrors play-tracker.html. Output: public/play-count-sheet.xlsx
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROSTER = [
    ("1", "Max Blatchford"), ("4", "Greyson Jones"), ("5", "Abrahm Tarwater"),
    ("7", "Desmond Grindal"), ("9", "Levi Hoskinson"), ("15", "Jaxon Duis"),
    ("16", "Obadiah Hurr"), ("17", "Ethan Lauver"), ("18", "Boone Clark"),
    ("20", "Easton Cook"), ("32", "Gideon Hawkinson"), ("42", "Nehemiah Hurr"),
    ("55", "Seamus Mason"), ("56", "Ethan Dinkel"), ("68", "Lee Napier"),
    ("72", "Logan Reihm"), ("75", "Levi Ketchem"), ("77", "Paddy Mason"),
    ("81", "Thomas Walker"), ("95", "Brayden Elliott"), ("98", "Declan Mason"),
]

PER_BLOCK = 6          # plays per rotation block
BLOCKS = 8             # rotation blocks
BLOCKS_PER_QTR = 2     # 2 blocks (12 plays) per quarter
PER_QTR = PER_BLOCK * BLOCKS_PER_QTR
TOTAL_BOXES = PER_BLOCK * BLOCKS         # 48
QUARTERS = BLOCKS // BLOCKS_PER_QTR      # 4
BLANK_ROWS = 3

RED = "BA1821"
INK = "191919"
GREY = "5A5A5A"
SHADE = "F1EDE8"
WHITE = "FFFFFF"

FIRST_BOX_COL = 4                        # col D
LAST_BOX_COL = FIRST_BOX_COL + TOTAL_BOXES - 1
TOTAL_COL = LAST_BOX_COL + 1

wb = Workbook()
ws = wb.active
ws.title = "Play Count"

thin = Side(style="thin", color="C9C2BB")
med = Side(style="medium", color=INK)
thick_red = Side(style="thick", color=RED)


def fill(hexc):
    return PatternFill("solid", fgColor=hexc)


def box_borders(col_index_zero_based):
    """Left border weight for a play box given its 0-based position."""
    if col_index_zero_based % PER_QTR == 0 and col_index_zero_based != 0:
        left = thick_red
    elif col_index_zero_based % PER_BLOCK == 0:
        left = med
    else:
        left = thin
    return Border(left=left, right=thin, top=thin, bottom=thin)


center = Alignment(horizontal="center", vertical="center")
left = Alignment(horizontal="left", vertical="center")

# --- Title / meta ---
ws["A1"] = "TCF SAINTS — PLAY COUNT & ROTATION SHEET"
ws["A1"].font = Font(bold=True, size=14, color=INK)
ws["A2"] = "Topeka Christian Saints · 2026"
ws["A2"].font = Font(size=9, color="555555")

ws["A3"] = "Opponent:"
ws["C3"] = "Date:"
ws["E3"] = "Unit (O / D / ST):"
for c in ("A3", "C3", "E3"):
    ws[c].font = Font(bold=True, size=10)

ws["A4"] = ("How to use: enter 1 in a box each play a player is on the field. "
            "Each block = 6 plays (rotate when a block fills; target 4-6 downs). "
            "Quarters Q1-Q4 shown by the red dividers. Total column auto-sums.")
ws["A4"].font = Font(size=9, italic=True, color="333333")

HDR_QTR = 6
HDR_ROT = 7
HDR_NUM = 8
FIRST_PLAYER = 9

# --- Quarter header row ---
for q in range(QUARTERS):
    start = FIRST_BOX_COL + q * PER_QTR
    end = start + PER_QTR - 1
    ws.merge_cells(start_row=HDR_QTR, start_column=start,
                   end_row=HDR_QTR, end_column=end)
    cell = ws.cell(row=HDR_QTR, column=start, value=f"Q{q + 1}")
    cell.fill = fill(RED)
    cell.font = Font(bold=True, color=WHITE, size=11)
    cell.alignment = center

# --- Rotation header row ---
for b in range(BLOCKS):
    start = FIRST_BOX_COL + b * PER_BLOCK
    end = start + PER_BLOCK - 1
    ws.merge_cells(start_row=HDR_ROT, start_column=start,
                   end_row=HDR_ROT, end_column=end)
    cell = ws.cell(row=HDR_ROT, column=start, value=f"ROT {b + 1}")
    cell.fill = fill(GREY)
    cell.font = Font(bold=True, color=WHITE, size=9)
    cell.alignment = center

# --- Number header row ---
head = {1: "#", 2: "Player", 3: "Pos"}
for col, label in head.items():
    cell = ws.cell(row=HDR_NUM, column=col, value=label)
    cell.fill = fill(INK)
    cell.font = Font(bold=True, color=WHITE, size=10)
    cell.alignment = left if col == 2 else center
    cell.border = Border(left=thin, right=thin, top=thin, bottom=med)

for i in range(TOTAL_BOXES):
    col = FIRST_BOX_COL + i
    cell = ws.cell(row=HDR_NUM, column=col, value=i + 1)
    cell.fill = fill(INK)
    cell.font = Font(bold=True, color=WHITE, size=8)
    cell.alignment = center
    b = box_borders(i)
    cell.border = Border(left=b.left, right=b.right, top=thin, bottom=med)

tcell = ws.cell(row=HDR_NUM, column=TOTAL_COL, value="Total")
tcell.fill = fill(RED)
tcell.font = Font(bold=True, color=WHITE, size=9)
tcell.alignment = center
tcell.border = Border(left=med, right=thin, top=thin, bottom=med)

# --- Player rows ---
rows = [(n, nm) for n, nm in ROSTER] + [("", "")] * BLANK_ROWS
for r, (num, name) in enumerate(rows):
    row = FIRST_PLAYER + r
    even = r % 2 == 1
    name_fill = fill("FAF8F6") if even else fill(WHITE)

    c = ws.cell(row=row, column=1, value=num)
    c.font = Font(bold=True, color=RED, size=12)
    c.alignment = center
    c.fill = name_fill
    c.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    c = ws.cell(row=row, column=2, value=name)
    c.font = Font(bold=True, size=11)
    c.alignment = left
    c.fill = name_fill
    c.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    c = ws.cell(row=row, column=3)
    c.alignment = center
    c.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for i in range(TOTAL_BOXES):
        col = FIRST_BOX_COL + i
        cell = ws.cell(row=row, column=col)
        block_even = (i // PER_BLOCK) % 2 == 0
        cell.fill = fill(WHITE) if block_even else fill(SHADE)
        cell.alignment = center
        cell.font = Font(size=9)
        cell.border = box_borders(i)

    # Total = sum of this row's play boxes
    first = f"{get_column_letter(FIRST_BOX_COL)}{row}"
    last = f"{get_column_letter(LAST_BOX_COL)}{row}"
    tc = ws.cell(row=row, column=TOTAL_COL, value=f"=SUM({first}:{last})")
    tc.font = Font(bold=True, size=11)
    tc.alignment = center
    tc.border = Border(left=med, right=thin, top=thin, bottom=thin)

# --- Column widths ---
ws.column_dimensions["A"].width = 4
ws.column_dimensions["B"].width = 22
ws.column_dimensions["C"].width = 6
for i in range(TOTAL_BOXES):
    ws.column_dimensions[get_column_letter(FIRST_BOX_COL + i)].width = 3.2
ws.column_dimensions[get_column_letter(TOTAL_COL)].width = 7

# Row heights
ws.row_dimensions[HDR_QTR].height = 18
ws.row_dimensions[HDR_ROT].height = 15
ws.row_dimensions[HDR_NUM].height = 15
for r in range(len(rows)):
    ws.row_dimensions[FIRST_PLAYER + r].height = 20

# --- Print setup: landscape, fit to one page wide ---
ws.page_setup.orientation = "landscape"
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.page_margins.left = ws.page_margins.right = 0.3
ws.page_margins.top = ws.page_margins.bottom = 0.4
ws.print_title_rows = f"{HDR_QTR}:{HDR_NUM}"
ws.freeze_panes = "D9"

import os
out = os.path.join(os.path.dirname(__file__), "..", "public", "play-count-sheet.xlsx")
out = os.path.abspath(out)
wb.save(out)
print("wrote", out)
